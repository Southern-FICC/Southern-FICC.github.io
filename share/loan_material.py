# -*- coding: utf-8 -*-
"""
资料分类汇总程序 - 单文件离线桌面版

运行方式：
    pip install -r requirements.txt
    python loan_material_app.py

打包方式：
    pyinstaller -F -w loan_material_app.py --name 资料分类汇总程序

说明：
    1. 程序不联网。
    2. 字段、下拉选项、上传资料清单、导出字段均从配置Excel读取。
    3. 模板/快照文件从用户选择的模板文件夹读取，按文件名前缀匹配，例如“3委托担保函模板.docx”。
    4. 附件只复制，不移动原文件。
"""

from __future__ import annotations

import os #路径处理及获取、文件存在性、文件夹创建、获取系统信息等
import re #正则表达式模块，提供文本（字符串）处理函数，如查找、替换、分割、文本格式规范化
import sys #提供python解释器与运行环境交互能力，处理在系统运行相关代码，如启动参数。
import shutil #os的补充，处理文件复制、移动、删除、打包和解压文件等
import traceback #提供错报信息链
from dataclasses import dataclass, field #负责表单数据管理。dataclass是该模块的装饰器，
from datetime import datetime, date #生成当前日期、文件日期、格式化日期，datetime（日期+事件）、date（日期）。常用于表单、统计和时间校验。
from pathlib import Path #路径处理模块，如：创建、拼接、遍历路径、获取文件名，核心类型是path。OS更偏向传统写法，操作更多是字符串拼接。
from typing import Any, Dict, List, Optional, Tuple #标注数据类型，以增强代码可读性和可维护性。Any包括字符串、整数、浮点数、布尔值、列表、字典、元组、对象实例，甚至函数、类、None。

try: #用于开始一个异常处理代码块
    from openpyxl import Workbook, load_workbook #workbook为该模块核心类，可以理解为整个xlsx文件的容器。！openpyxl不是标准库，因此没有详细提示。
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side #styles是该模块的专门负责excel样式设置的子模块。
    from openpyxl.utils import get_column_letter #子模块，处理列坐标、列号转字母列名等，以定位列进行处理数据。
except Exception as exc:  # Except代码块，Exception 是Python中所有常规异常的基类，如果不写Exception的话，就变成裸异常捕获，会捕获所有异常，包括键盘按键冲突、系统关闭等非程序异常，写的话，而且保存为as，方便输出日志。
    print("缺少 openpyxl，请先安装：pip install openpyxl")
    raise #用于重新抛出异常，如果外层有额外的try/except的话，就继续，没有就报错结束。
# JSON里面路径可以用"/"，或者"\\"，如果是"\"，单个 \ 会被当作转义符
try:
    from PySide6.QtCore import Qt, QUrl, Signal 
    from PySide6.QtGui import QDesktopServices, QDoubleValidator, QIntValidator, QFont
    from PySide6.QtWidgets import (
        QApplication,
        QMainWindow,
        QWidget,
        QStackedWidget,
        QVBoxLayout,
        QHBoxLayout,
        QGridLayout,
        QLabel,
        QPushButton,
        QFileDialog,
        QLineEdit,
        QComboBox,
        QGroupBox,
        QScrollArea,
        QMessageBox,
        QFrame,
        QProgressDialog,
        QSizePolicy,
        QTextEdit,
    )
except Exception as exc:  # pragma: no cover
    print("缺少 PySide6，请先安装：pip install PySide6")
    raise


# =========================
# 基础工具函数
# =========================

def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == "" #判断这个值是不是空的，且return后面接表达式，不能接value=””赋值语句。
# str(value)把value转换成一个字符串对象，后面的strip()是在字符串对象上调用方法。 类负责定义方法，对象负责调用方法；实例方法会自动拿到这个对象自己作为 self。
# __init__为方法/构造函数。self不是构造函数，是被构造出来的实例/对象。str(value)，value就是self，str是字符串类，s = str(value) ，这里 value 是原料，s 是产出的实例（相当于 self）。str(value).strip(),strip是str类的方法，此处为字符串对象调用的方法strip()，该对象自动传入strip里面。
# ctrl点击类、对象、方法后出现的pyi文件叫类型桩文件，用来给类型检查器(如mypy、pyright)看的。
# Python 有个统一规则：方法调用时，点号左边的那个东西，会被自动当成第一个参数传进去。

def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer(): #isinstance用于是python内置函数，用于判断变量类型，（value,float），判断value是否为float或其子类。相比于type(obj)==someclass，isinstance更推荐，因为会把子类也纳入判断，不仅仅父类。
        return str(int(value))
    return str(value)


def clean_text(value: Any) -> str:
    """导出前统一清洗：去前后空格、英文括号转中文括号。"""
    text = as_text(value).strip()
    text = text.replace("(", "（").replace(")", "）")
    return text


def safe_filename(value: Any, default: str = "未填写") -> str:
    text = clean_text(value) or default
    text = re.sub(r'[\\/:*?"<>|]+', "_", text) #repl是replacement的缩写，这里对应用到的参数为pattern（要匹配的正则），repl是用于替换的字符串，string是需要被替换的字符串。[]是正则里的”字符集合“，意思是匹配里面任意一个字符。r‘’是原始字符串写法，主要是为了让反斜杠\\按字面意义参与正则表达式，而不是被转义。+号表示这些非法字符如果连续出现，会被一次性替换成一个_
    text = re.sub(r"\s+", " ", text).strip() #r''和r""都一样，都是让\不被转义。 \s在正则中表示任意空白字符，包括空格、制表符\t、换行\n、回车\r，这里\s+表示连续多个空白字符替换成一个普通空格。
    return text or default 


def normalize_date(value: Any) -> str: #Any包括字符串、整数、浮点数、布尔值、列表、字典、元组、对象实例，甚至函数、类、None。
    """
    兼容常见日期输入：
    20260708、2026-07-08、2026.7.8、2026/7/8、2026年7月8日。
    返回 YYYY/MM/DD；无法识别则返回清洗后的原值。
    """
    if value is None:
        return ""
    if isinstance(value, datetime): #因为datetime是date的子类，所以需要先判断子类。用 isinstance() 检查对象是否属于某个类或其子类。
        return value.strftime("%Y/%m/%d") #strftime的英文全称通常解释为string format time，把时间格式按照%Y/%m/%d格式转为字符串。"%Y/%m/%d"表示四位年份、两位月份、两位日期
    if isinstance(value, date): 
        return value.strftime("%Y/%m/%d") #可以写成elif isinstance(value, date): ，然后else: {空行} text = clean_text(value)。。。但是因为如果执行了return，其实后面也不会继续，写else反而会复杂，这样更容易阅读。

    text = clean_text(value)
    if not text: #如果写成if text is not True，表示的是text这个对象是否不是布尔值True本身。
        return ""

    text2 = text.replace("年", "/").replace("月", "/").replace("日", "")
    text2 = text2.replace(".", "/").replace("-", "/")
    text2 = re.sub(r"\s+", "", text2)
    # date的标准文本格式为2026-07-12，datetime的标准文本格式为2026-07-12 15:30:45
    # 20260708
    if re.fullmatch(r"\d{8}", text2): #\d是正则表达式，代表任意一个数字字符，{8}表示前面的规则重复8次，所以是匹配8位数的字符串。
        try:
            dt = datetime.strptime(text2, "%Y%m%d") # strptime通常解释为 string parse time，也就是“把字符串解析成时间”，将格式为%Y%m%d的字符串text2，解析成datetime对象。
            return dt.strftime("%Y/%m/%d") #这里是把dt转为格式为%Y/%m/%d的字符串。
        except ValueError:
            return text  #如果发生错误，返回原值。

    # 2026/7/8 或 2026/07/08
    m = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", text2) #\d{1,2}是正则表达式，{1,2}表示重复1次或2次。这里的/是普通斜杠，从而匹配2026/7/8或者2026/07/08这两种格式。这里的m就是日期字符串本身，但是包在Match对象里面，包括三类东西，一个是整段匹配到的文本，可以用m.group(0)取出来，二是正则捕获到的部分，三是匹配发生的位置、是否成功等。Python不直接把这个值裸返回，而是放进Match对象，方便之后取整段、取分组、去位置。
    if m:
        y, mo, d = map(int, m.groups()) #m.groups只取捕获的部分，只取正则表达式部分，所以斜杠/不在内。m.groups没有参数，是一次性获取所有捕获部分返回元组tuple，然后传入int()，处理成整数，然后逐一赋值，所以没有.m.groups(1)。
        try:                                              #如果是m.group(0)，没有s的，那就会把/也放入，取的是完整文本，返回的是字符串类型。另外，group(0)返回整段字符串2026/7/8，group(1)返回第一个括号捕捉到的字符串2026，group(2)返回的是7，group(1,2,3)返回元组。
            dt = datetime(y, mo, d) #直接有年月日值，不用通过strptime按照字符串的格式解析为datetime。
            return dt.strftime("%Y/%m/%d")
        except ValueError:
            return text

    # Excel序列号粗略兼容，不鼓励直接输入
    if re.fullmatch(r"\d{5}", text2):
        try:
            from openpyxl.utils.datetime import from_excel #utils是openpyxl模块下的常用子模块，英文全称为utilities，工具，专门放”杂项辅助函数“。datetime是utils的子模块，放日期相关工具函数。
            dt = from_excel(int(text2)) #from_excel是把日期序列号转为日期，to_excel就是把日期转为序列号。
            return dt.strftime("%Y/%m/%d")
        except Exception:
            pass

    return text


def extract_month(value: Any) -> str:
    dt = normalize_date(value)
    m = re.fullmatch(r"(\d{4})/(\d{2})/(\d{2})", dt)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    return ""


def to_number(value: Any) -> Optional[float]: #表示返回值要么是float，要么是None，即float|None。Optional只能接受一个类型参数。如果想要多个类型参数，需要嵌套Union，如：Optioinal[Union[float,int]]。不过，更推荐写成def to_number(value: Any) -> float | int | None:
    text = clean_text(value).replace(",", "")        #from typing import Optional, Union     
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def is_compressed_file(path: Path) -> bool:
    return path.suffix.lower() in {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz"}
# path的标准格式有相对路径，如："files/资料.zip"，以及绝对路径，如：r"G:\000职业发展\资料.zip"，r表示原始字符串，避免\n、\t、\d、\s等被转义，或者"G:/000职业发展/资料.zip"，windows也支持正斜杠。
# suffix是Path对象的属性，表示文件的最后一个扩展名，并且包含开头的点。另外，多重扩展名，如："backup.tar.gz"，只会返回最后一个。如果用suffixes，就会都返回，如：[".tar", ".gz"]。
# Path和PurePath都是模块Pathlib的类，而suffix是PurePath类的属性@property，不过定义Path类的时候设置PurePath为参数，所以Path为PurePath的子类，class Path(PurePath):，所以Path可以使用PurePath中定义的属性和方法。

def open_file(path: Path) -> None:
    if path and path.exists():
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(path.resolve())))


def find_template_file(template_dir: Path, prefix: Any) -> Optional[Path]:
    if is_blank(prefix) or not template_dir or not template_dir.exists():
        return None
    pfx = str(prefix).strip()
    candidates = sorted([p for p in template_dir.iterdir() if p.is_file() and p.name.startswith(pfx)])
    return candidates[0] if candidates else None


# =========================
# 配置读取
# =========================

@dataclass
class FieldDef:
    path_name: str
    area: str
    row_no: int
    order: int
    key: str
    label: str
    widget_type: str
    required: bool = False
    default: str = ""
    data_source: str = ""
    read_only: bool = False
    disabled: bool = False
    unit: str = ""
    validation: str = ""
    placeholder: str = ""
    action: str = ""
    add_limit: int = 0
    export_field: str = ""
    note: str = ""


@dataclass
class UploadItem:
    row_no: int
    name: str
    path_scope: str = "全部"
    required: bool = False
    allow_multi: bool = True
    forbid_folder: bool = True
    forbid_zip: bool = True
    template_label: str = ""
    template_type: str = ""
    template_prefix: str = ""
    note: str = ""


@dataclass
class ExportCol:
    order: int
    col_name: str
    source_key: str = ""
    default: str = ""
    data_type: str = "文本"
    excel_format: str = "@"
    rule: str = ""
    note: str = ""


@dataclass
class AppConfig:
    config_path: Path
    program_info: Dict[str, str] = field(default_factory=dict)
    fields: List[FieldDef] = field(default_factory=list)
    options: Dict[str, List[str]] = field(default_factory=dict)
    manager_to_branch: Dict[str, str] = field(default_factory=dict)
    guarantee_to_contract: Dict[str, str] = field(default_factory=dict)
    upload_items: List[UploadItem] = field(default_factory=list)
    export_cols: List[ExportCol] = field(default_factory=list)

    @property
    def program_name(self) -> str:
        value = self.program_info.get("程序名称") or self.config_path.stem
        return safe_filename(value, default=self.config_path.stem)


def read_worksheet_dicts(wb, sheet_name: str) -> List[Dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [as_text(c.value).strip() for c in ws[1]]
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(is_blank(x) for x in row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        rows.append(item)
    return rows


def read_config(config_path: Path) -> AppConfig:
    wb = load_workbook(config_path, data_only=True)
    cfg = AppConfig(config_path=config_path)

    # 程序信息
    for row in read_worksheet_dicts(wb, "程序信息"):
        k = clean_text(row.get("配置项"))
        v = clean_text(row.get("配置值"))
        if k:
            cfg.program_info[k] = v

    # 页面字段
    fields: List[FieldDef] = []
    for row in read_worksheet_dicts(wb, "页面字段"):
        try:
            fd = FieldDef(
                path_name=clean_text(row.get("路径")),
                area=clean_text(row.get("区域")),
                row_no=int(row.get("行号") or 1),
                order=int(row.get("排序") or 0),
                key=clean_text(row.get("字段键")),
                label=clean_text(row.get("字段名称")),
                widget_type=clean_text(row.get("控件类型")),
                required=clean_text(row.get("是否必填")) == "是",
                default=clean_text(row.get("默认值")),
                data_source=clean_text(row.get("数据源键")),
                read_only=clean_text(row.get("只读")) == "是",
                disabled=clean_text(row.get("禁用/标灰")) == "是",
                unit=clean_text(row.get("单位")),
                validation=clean_text(row.get("校验规则")),
                placeholder=clean_text(row.get("占位提示")),
                action=clean_text(row.get("关联规则/动作")),
                add_limit=int(row.get("新增上限") or 0),
                export_field=clean_text(row.get("导出字段")),
                note=clean_text(row.get("备注")),
            )
            if fd.path_name and fd.key:
                fields.append(fd)
        except Exception:
            continue
    cfg.fields = sorted(fields, key=lambda x: (x.path_name, x.area, x.row_no, x.order))

    # 下拉选项
    option_rows = read_worksheet_dicts(wb, "下拉选项")
    temp: Dict[str, List[Tuple[int, str]]] = {}
    for row in option_rows:
        enabled = clean_text(row.get("启用"))
        if enabled and enabled != "是":
            continue
        key = clean_text(row.get("数据源键"))
        name = clean_text(row.get("选项名称"))
        order = int(row.get("排序") or 0)
        if key and name:
            temp.setdefault(key, []).append((order, name))
    for key, items in temp.items():
        cfg.options[key] = [x[1] for x in sorted(items, key=lambda t: t[0])]

    # 下拉映射
    for row in read_worksheet_dicts(wb, "下拉映射"):
        map_type = clean_text(row.get("映射类型"))
        parent_value = clean_text(row.get("父级选项名称"))
        child_key = clean_text(row.get("子级字段键/数据源键"))
        child_value = clean_text(row.get("子级值"))
        if not parent_value:
            continue
        if "客户经理" in map_type and (child_key == "branch_name" or "支行" in child_key):
            cfg.manager_to_branch[parent_value] = child_value
        elif "担保" in map_type or child_key == "guarantee_contract_no":
            cfg.guarantee_to_contract[parent_value] = child_value

    # 上传资料清单
    uploads: List[UploadItem] = []
    for row in read_worksheet_dicts(wb, "上传资料清单"):
        try:
            uploads.append(
                UploadItem(
                    row_no=int(row.get("行号") or 0),
                    name=clean_text(row.get("资料名称/显示文本")),
                    path_scope=clean_text(row.get("适用路径")) or "全部",
                    required=clean_text(row.get("是否必传")) == "是",
                    allow_multi=clean_text(row.get("允许多文件")) != "否",
                    forbid_folder=clean_text(row.get("禁止文件夹")) != "否",
                    forbid_zip=clean_text(row.get("禁止压缩包")) != "否",
                    template_label=clean_text(row.get("模板标签")),
                    template_type=clean_text(row.get("模板类型")),
                    template_prefix=clean_text(row.get("模板文件名前缀")),
                    note=clean_text(row.get("上传说明")),
                )
            )
        except Exception:
            continue
    cfg.upload_items = sorted([u for u in uploads if u.row_no and u.name], key=lambda x: x.row_no)

    # 导出字段
    cols: List[ExportCol] = []
    for row in read_worksheet_dicts(wb, "导出字段"):
        try:
            cols.append(
                ExportCol(
                    order=int(row.get("排序") or 0),
                    col_name=clean_text(row.get("导出列名")),
                    source_key=clean_text(row.get("来源字段键")),
                    default=clean_text(row.get("默认值")),
                    data_type=clean_text(row.get("数据类型")) or "文本",
                    excel_format=clean_text(row.get("Excel格式")) or "@",
                    rule=clean_text(row.get("清洗/转换规则")),
                    note=clean_text(row.get("备注")),
                )
            )
        except Exception:
            continue
    cfg.export_cols = sorted([c for c in cols if c.col_name], key=lambda x: x.order)

    return cfg


# =========================
# 控件封装
# =========================

class NoWheelComboBox(QComboBox):
    def wheelEvent(self, event):  # noqa: N802
        event.ignore()

class FieldBox(QWidget):
    """一个字段：标签 + 输入控件 + 单位/计数/按钮。"""

    def __init__(self, field_def: FieldDef, cfg: AppConfig, parent_form: "FormPage", dynamic_index: Optional[int] = None):
        super().__init__()
        self.fd = field_def
        self.cfg = cfg
        self.parent_form = parent_form
        self.dynamic_index = dynamic_index
        self.count_label: Optional[QLabel] = None
        self.widget: QWidget

        outer = QVBoxLayout(self)
        outer.setContentsMargins(4, 4, 4, 4)
        outer.setSpacing(3)

        title = QLabel(self.fd.label + (" *" if self.fd.required else ""))
        title.setWordWrap(True)
        title.setStyleSheet("font-size: 13px; font-weight: 600; color: #222;")
        outer.addWidget(title)

        line = QHBoxLayout()
        line.setContentsMargins(0, 0, 0, 0)
        line.setSpacing(4)
        outer.addLayout(line)

        self.widget = self._create_widget()
        line.addWidget(self.widget, 1)

        if self.fd.unit:
            unit = QLabel(self.fd.unit)
            unit.setStyleSheet("font-size: 12px; color: #333;")
            line.addWidget(unit)

        if self._needs_count_label():
            self.count_label = QLabel("0")
            self.count_label.setFixedWidth(34)
            self.count_label.setAlignment(Qt.AlignCenter)
            self.count_label.setStyleSheet("font-size: 12px; color: #555;")
            line.addWidget(self.count_label)
            if isinstance(self.widget, QLineEdit):
                self.widget.textChanged.connect(self._refresh_length_warning)
                self._refresh_length_warning(self.widget.text())

        if self.fd.key == "entrusted_guarantee_letter_no":
            btn = QPushButton("复制借款合同号")
            btn.setCursor(Qt.PointingHandCursor)
            btn.setStyleSheet("font-size: 12px; padding: 4px 8px;")
            btn.clicked.connect(self._copy_loan_contract_no)
            line.addWidget(btn)

        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

    def _create_widget(self) -> QWidget:
        wt = self.fd.widget_type
        disabled = self.fd.disabled or self.fd.read_only

        if wt == "下拉框":
            cb = NoWheelComboBox()
            cb.setEditable(False)
            cb.addItem("")
            for item in self.cfg.options.get(self.fd.data_source, []):
                cb.addItem(item)
            if self.fd.default:
                idx = cb.findText(self.fd.default)
                if idx >= 0:
                    cb.setCurrentIndex(idx)
            cb.setEnabled(not disabled)
            self._apply_base_style(cb, disabled)
            return cb

        # 自动填充框、文本框、数字框、日期框均用QLineEdit，便于配置驱动
        le = QLineEdit()
        le.setText(self.fd.default or "")
        if self.fd.placeholder:
            le.setPlaceholderText(self.fd.placeholder)
        if wt == "数字框":
            validator = QDoubleValidator(0, 999999999999, 6, le)
            validator.setNotation(QDoubleValidator.StandardNotation)
            le.setValidator(validator)
        elif wt == "日期框":
            le.editingFinished.connect(self._normalize_date_on_finish)
        le.setReadOnly(disabled)
        le.setEnabled(True)  # readOnly也保持可复制文字
        self._apply_base_style(le, disabled)
        return le

    def _apply_base_style(self, widget: QWidget, disabled: bool = False, invalid: bool = False) -> None:
        if invalid:
            style = "background: #ffe5e5; border: 1px solid #cc0000; border-radius: 4px; padding: 5px; font-size: 13px;"
        elif disabled:
            style = "background: #e9e9e9; border: 1px solid #aaa; border-radius: 4px; padding: 5px; font-size: 13px; color: #666;"
        else:
            style = "background: #ffffff; border: 1px solid #999; border-radius: 4px; padding: 5px; font-size: 13px;"
        widget.setStyleSheet(style)

    def _needs_count_label(self) -> bool:
        return "证件号码" in self.fd.label or self.fd.key.endswith("id_no") or self.fd.key.endswith("_id_no")

    def _refresh_length_warning(self, text: str) -> None:
        length = len(text.strip())
        if self.count_label:
            self.count_label.setText(str(length))
        invalid = length != 18
        if isinstance(self.widget, QLineEdit):
            self._apply_base_style(self.widget, self.fd.disabled or self.fd.read_only, invalid=invalid)

    def _normalize_date_on_finish(self) -> None:
        if not isinstance(self.widget, QLineEdit):
            return
        value = self.widget.text().strip()
        if not value:
            self._apply_base_style(self.widget, self.fd.disabled or self.fd.read_only)
            return
        new_value = normalize_date(value)
        self.widget.setText(new_value)
        invalid = not bool(re.fullmatch(r"\d{4}/\d{2}/\d{2}", new_value))
        self._apply_base_style(self.widget, self.fd.disabled or self.fd.read_only, invalid=invalid)

    def _copy_loan_contract_no(self) -> None:
        value = self.parent_form.get_field_value("loan_contract_no")
        self.set_value(value)

    def value(self) -> str:
        if isinstance(self.widget, QComboBox):
            return clean_text(self.widget.currentText())
        if isinstance(self.widget, QLineEdit):
            if self.fd.widget_type == "日期框":
                return normalize_date(self.widget.text())
            return clean_text(self.widget.text())
        return ""

    def set_value(self, value: Any) -> None:
        value = clean_text(value)
        if isinstance(self.widget, QComboBox):
            idx = self.widget.findText(value)
            if idx < 0 and value:
                self.widget.addItem(value)
                idx = self.widget.findText(value)
            if idx >= 0:
                self.widget.setCurrentIndex(idx)
        elif isinstance(self.widget, QLineEdit):
            self.widget.setText(value)


class DropArea(QFrame):
    files_changed = Signal()

    def __init__(self, item: UploadItem):
        super().__init__()
        self.item = item
        self.files: List[Path] = []
        self.setAcceptDrops(True)
        self.setMinimumHeight(110)
        self.setStyleSheet(
            "QFrame { background: #ffffff; border: 2px dashed #888; border-radius: 10px; }"
            "QLabel { border: none; background: transparent; }"
            "QPushButton { border: 1px solid #888; background: #f7f7f7; padding: 4px 8px; border-radius: 4px; }"
        )

        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(8, 8, 8, 8)
        self.layout.setSpacing(4)

        self.plus = QLabel("+")
        self.plus.setAlignment(Qt.AlignCenter)
        self.plus.setStyleSheet("font-size: 38px; font-weight: 300; color: #555;")
        self.layout.addWidget(self.plus)

        self.hint = QLabel("上传附件")
        self.hint.setAlignment(Qt.AlignCenter)
        self.hint.setStyleSheet("font-size: 12px; color: #555;")
        self.layout.addWidget(self.hint)

        self.file_list_label = QLabel("")
        self.file_list_label.setWordWrap(True)
        self.file_list_label.setStyleSheet("font-size: 12px; color: #222; line-height: 1.3;")
        self.layout.addWidget(self.file_list_label)

        btn_line = QHBoxLayout()
        btn_line.setContentsMargins(0, 0, 0, 0)
        self.select_btn = QPushButton("选择文件")
        self.clear_btn = QPushButton("清空")
        self.select_btn.clicked.connect(self.select_files)
        self.clear_btn.clicked.connect(self.clear_files)
        btn_line.addStretch(1)
        btn_line.addWidget(self.select_btn)
        btn_line.addWidget(self.clear_btn)
        btn_line.addStretch(1)
        self.layout.addLayout(btn_line)

    def mousePressEvent(self, event):  # noqa: N802
        if event.button() == Qt.LeftButton:
            self.select_files()
        super().mousePressEvent(event)

    def dragEnterEvent(self, event):  # noqa: N802
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):  # noqa: N802
        paths = []
        for url in event.mimeData().urls():
            if url.isLocalFile():
                paths.append(Path(url.toLocalFile()))
        if paths:
            self.add_files(paths)
        event.acceptProposedAction()

    def select_files(self) -> None:
        file_names, _ = QFileDialog.getOpenFileNames(self, "选择上传附件", "", "所有文件 (*.*)")
        if file_names:
            self.add_files([Path(x) for x in file_names])

    def clear_files(self) -> None:
        self.files.clear()
        self.refresh()
        self.files_changed.emit()

    def add_files(self, paths: List[Path]) -> None:
        rejected: List[str] = []
        for path in paths:
            if self.item.forbid_folder and path.is_dir():
                rejected.append(f"{path.name}：不能上传文件夹")
                continue
            if self.item.forbid_zip and is_compressed_file(path):
                rejected.append(f"{path.name}：不能上传压缩文件")
                continue
            if not path.exists() or not path.is_file():
                rejected.append(f"{path.name}：文件不存在或不可读取")
                continue
            if (not self.item.allow_multi) and self.files:
                self.files = [path]
            elif path not in self.files:
                self.files.append(path)
        self.refresh()
        self.files_changed.emit()
        if rejected:
            QMessageBox.warning(self, "部分文件未添加", "\n".join(rejected))

    def refresh(self) -> None:
        if not self.files:
            self.plus.show()
            self.hint.show()
            self.file_list_label.setText("")
            return
        self.plus.hide()
        self.hint.setText(f"已上传 {len(self.files)} 个文件，点击可继续添加")
        self.hint.show()
        lines = [f"• {p.name}" for p in self.files]
        self.file_list_label.setText("\n".join(lines))


# =========================
# 页面：启动配置选择
# =========================

class StartupPage(QWidget):
    loaded = Signal(object, object)  # cfg, template_dir

    def __init__(self):
        super().__init__()
        self.setStyleSheet("background: #E2F0D9;")
        root = QVBoxLayout(self)
        root.setContentsMargins(40, 40, 40, 40)
        root.setSpacing(16)

        title = QLabel("资料分类汇总程序")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 24px; font-weight: 700; color: #1f1f1f;")
        root.addWidget(title)

        subtitle = QLabel("请选择配置Excel和模板文件夹。程序全程本地运行，不联网。")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("font-size: 14px; color: #333;")
        root.addWidget(subtitle)

        form = QGridLayout()
        form.setHorizontalSpacing(8)
        form.setVerticalSpacing(12)
        root.addLayout(form)

        self.config_edit = QLineEdit()
        self.template_edit = QLineEdit()
        self.config_edit.setPlaceholderText("选择配置Excel，例如：资料分类汇总程序配置表_试验版.xlsx")
        self.template_edit.setPlaceholderText("选择模板文件夹，例如：templates")
        self._style_edit(self.config_edit)
        self._style_edit(self.template_edit)

        default_cfg = Path.cwd() / "资料分类汇总程序配置表_试验版.xlsx"
        if not default_cfg.exists():
            default_cfg = Path(__file__).with_name("资料分类汇总程序配置表_试验版.xlsx")
        if default_cfg.exists():
            self.config_edit.setText(str(default_cfg))

        cfg_btn = QPushButton("浏览")
        tpl_btn = QPushButton("浏览")
        cfg_btn.clicked.connect(self.browse_config)
        tpl_btn.clicked.connect(self.browse_template_dir)

        form.addWidget(QLabel("配置Excel："), 0, 0)
        form.addWidget(self.config_edit, 0, 1)
        form.addWidget(cfg_btn, 0, 2)
        form.addWidget(QLabel("模板文件夹："), 1, 0)
        form.addWidget(self.template_edit, 1, 1)
        form.addWidget(tpl_btn, 1, 2)

        note = QLabel("模板文件可为空；为空时模板/模板快照链接点击无反应。")
        note.setStyleSheet("font-size: 12px; color: #555;")
        root.addWidget(note)

        start_btn = QPushButton("启动程序")
        start_btn.setFixedHeight(44)
        start_btn.setCursor(Qt.PointingHandCursor)
        start_btn.setStyleSheet("font-size: 16px; font-weight: 700; background: #2f6f3e; color: white; border-radius: 6px;")
        start_btn.clicked.connect(self.load_config)
        root.addWidget(start_btn)
        root.addStretch(1)

    def _style_edit(self, edit: QLineEdit) -> None:
        edit.setStyleSheet("background: white; border: 1px solid #888; border-radius: 4px; padding: 7px; font-size: 13px;")

    def browse_config(self) -> None:
        file_name, _ = QFileDialog.getOpenFileName(self, "选择配置Excel", "", "Excel文件 (*.xlsx *.xlsm)")
        if file_name:
            self.config_edit.setText(file_name)

    def browse_template_dir(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "选择模板文件夹")
        if folder:
            self.template_edit.setText(folder)

    def load_config(self) -> None:
        cfg_path = Path(self.config_edit.text().strip())
        tpl_text = self.template_edit.text().strip()
        template_dir = Path(tpl_text) if tpl_text else Path("")
        if not cfg_path.exists():
            QMessageBox.warning(self, "配置文件不存在", "请选择有效的配置Excel文件。")
            return
        try:
            cfg = read_config(cfg_path)
            self.loaded.emit(cfg, template_dir)
        except Exception as exc:
            QMessageBox.critical(self, "读取配置失败", f"读取配置Excel失败：\n{exc}\n\n{traceback.format_exc()}")


# =========================
# 页面：路径选择
# =========================

class PathSelectPage(QWidget):
    path_selected = Signal(str)

    def __init__(self, cfg: AppConfig):
        super().__init__()
        self.cfg = cfg
        self.setStyleSheet("background: #E2F0D9;")
        root = QVBoxLayout(self)
        root.setContentsMargins(60, 50, 60, 50)
        root.setSpacing(24)

        title = QLabel(cfg.program_name)
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 24px; font-weight: 700; color: #1f1f1f;")
        root.addWidget(title)

        prompt = QLabel("请选择借款人路径")
        prompt.setAlignment(Qt.AlignCenter)
        prompt.setStyleSheet("font-size: 16px; color: #333;")
        root.addWidget(prompt)

        paths = []
        for fd in cfg.fields:
            if fd.path_name and fd.path_name not in paths:
                paths.append(fd.path_name)

        # 按用户需求固定顺序：A在第一行，B在第二行
        paths = sorted(paths, key=lambda x: 0 if "路径A" in x else 1)
        for path_name in paths:
            btn = QPushButton(path_name)
            btn.setMinimumHeight(90)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setStyleSheet(
                "QPushButton { background: #ffffff; border: 2px solid #355e3b; border-radius: 12px; "
                "font-size: 20px; font-weight: 700; color: #1f1f1f; }"
                "QPushButton:hover { background: #f2fff2; }"
            )
            btn.clicked.connect(lambda checked=False, p=path_name: self.path_selected.emit(p))
            root.addWidget(btn)

        root.addStretch(1)


# =========================
# 页面：信息填写
# =========================

class FormPage(QWidget):
    submitted = Signal(dict, list, str)  # form_data, co_borrowers, selected_path
    back_requested = Signal()

    def __init__(self, cfg: AppConfig, selected_path: str):
        super().__init__()
        self.cfg = cfg
        self.selected_path = selected_path
        self.field_boxes: Dict[str, FieldBox] = {}
        self.co_borrower_field_defs: List[FieldDef] = []
        self.co_borrower_boxes: List[Dict[str, FieldBox]] = []
        self.co_limit = 3

        self.setStyleSheet("background: #E2F0D9;")
        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        top = QHBoxLayout()
        back_btn = QPushButton("返回路径选择")
        back_btn.clicked.connect(self.back_requested.emit)
        back_btn.setStyleSheet("font-size: 13px; padding: 6px 10px;")
        title = QLabel(selected_path)
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: 700; color: #1f1f1f;")
        top.addWidget(back_btn)
        top.addWidget(title, 1)
        top.addSpacing(100)
        root.addLayout(top)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: #E2F0D9; }")
        body = QWidget()
        body.setStyleSheet("background: #E2F0D9;")
        self.body_layout = QVBoxLayout(body)
        self.body_layout.setContentsMargins(4, 4, 4, 4)
        self.body_layout.setSpacing(12)
        scroll.setWidget(body)
        root.addWidget(scroll, 1)

        self._build_form()

        bottom = QHBoxLayout()
        bottom.addStretch(1)
        submit_btn = QPushButton("提交资料")
        submit_btn.setFixedHeight(42)
        submit_btn.setMinimumWidth(150)
        submit_btn.setCursor(Qt.PointingHandCursor)
        submit_btn.setStyleSheet("font-size: 16px; font-weight: 700; background: #2f6f3e; color: white; border-radius: 6px;")
        submit_btn.clicked.connect(self.submit)
        bottom.addWidget(submit_btn)
        root.addLayout(bottom)

        self._connect_linkages()

    def _path_fields(self) -> List[FieldDef]:
        return [f for f in self.cfg.fields if f.path_name == self.selected_path]

    def _build_form(self) -> None:
        fields = self._path_fields()
        normal_fields = []
        for f in fields:
            if f.area == "共同借款人基本信息":
                self.co_borrower_field_defs.append(f)
                if f.add_limit:
                    self.co_limit = f.add_limit
            elif f.key == "co_borrower_add":
                if f.add_limit:
                    self.co_limit = f.add_limit
                # 加号按钮放在“共同借款人”区域里，不作为普通字段
            else:
                normal_fields.append(f)

        area_order: List[str] = []
        for f in normal_fields:
            if f.area not in area_order:
                area_order.append(f.area)

        for area in area_order:
            group = QGroupBox(area)
            group.setStyleSheet(
                "QGroupBox { background: rgba(255,255,255,0.45); border: 1px solid #9ab88d; border-radius: 8px; "
                "margin-top: 14px; padding: 12px; font-size: 15px; font-weight: 700; }"
                "QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 4px; }"
            )
            grid = QGridLayout(group)
            grid.setHorizontalSpacing(10)
            grid.setVerticalSpacing(10)

            area_fields = [f for f in normal_fields if f.area == area]
            # 既尊重配置里的行号，也在一行字段过多时自动换行，避免界面太宽
            row_groups: Dict[int, List[FieldDef]] = {}
            for f in area_fields:
                row_groups.setdefault(f.row_no, []).append(f)

            grid_row = 0
            max_cols = 4
            for row_no in sorted(row_groups):
                row_fields = sorted(row_groups[row_no], key=lambda x: x.order)
                for idx, fd in enumerate(row_fields):
                    r = grid_row + idx // max_cols
                    c = idx % max_cols
                    fb = FieldBox(fd, self.cfg, self)
                    self.field_boxes[fd.key] = fb
                    grid.addWidget(fb, r, c)
                grid_row += (len(row_fields) + max_cols - 1) // max_cols

            self.body_layout.addWidget(group)

            # 在借款人基本信息后插入共同借款人动态区域
            if area == "借款人基本信息":
                self._build_co_borrower_area()

        self.body_layout.addStretch(1)

    def _build_co_borrower_area(self) -> None:
        self.co_group = QGroupBox("共同借款人基本信息")
        self.co_group.setStyleSheet(
            "QGroupBox { background: rgba(255,255,255,0.35); border: 1px dashed #789c6f; border-radius: 8px; "
            "margin-top: 14px; padding: 12px; font-size: 15px; font-weight: 700; }"
            "QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 4px; }"
        )
        self.co_layout = QVBoxLayout(self.co_group)
        self.co_layout.setContentsMargins(8, 12, 8, 8)
        self.co_layout.setSpacing(8)

        btn_line = QHBoxLayout()
        self.add_co_btn = QPushButton("＋ 新增共同借款人")
        self.add_co_btn.setCursor(Qt.PointingHandCursor)
        self.add_co_btn.setStyleSheet("font-size: 14px; padding: 7px 12px; background: white; border: 1px solid #777; border-radius: 5px;")
        self.add_co_btn.clicked.connect(self.add_co_borrower)
        btn_line.addWidget(self.add_co_btn)
        btn_line.addStretch(1)
        self.co_layout.addLayout(btn_line)

        self.co_rows_container = QVBoxLayout()
        self.co_layout.addLayout(self.co_rows_container)
        self.body_layout.addWidget(self.co_group)

    def add_co_borrower(self) -> None:
        if len(self.co_borrower_boxes) >= self.co_limit:
            QMessageBox.information(self, "已达上限", f"共同借款人最多新增 {self.co_limit} 位。")
            return
        idx = len(self.co_borrower_boxes) + 1
        row_group = QGroupBox(f"共同借款人 {idx}")
        row_group.setStyleSheet(
            "QGroupBox { background: #f9fff9; border: 1px solid #bbb; border-radius: 6px; margin-top: 10px; padding: 10px; }"
            "QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 4px; }"
        )
        grid = QGridLayout(row_group)
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(8)
        boxes: Dict[str, FieldBox] = {}
        for col, fd in enumerate(sorted(self.co_borrower_field_defs, key=lambda x: x.order)):
            fb = FieldBox(fd, self.cfg, self, dynamic_index=idx)
            boxes[fd.key] = fb
            grid.addWidget(fb, 0, col)
        self.co_borrower_boxes.append(boxes)
        self.co_rows_container.addWidget(row_group)
        if len(self.co_borrower_boxes) >= self.co_limit:
            self.add_co_btn.setEnabled(False)
            self.add_co_btn.setText(f"已达到上限：{self.co_limit} 位")

    def _connect_linkages(self) -> None:
        manager = self.field_boxes.get("manager_name")
        branch = self.field_boxes.get("branch_name")
        if manager and branch and isinstance(manager.widget, QComboBox):
            manager.widget.currentTextChanged.connect(self._on_manager_changed)

        org = self.field_boxes.get("guarantee_org")
        contract = self.field_boxes.get("guarantee_contract_no")
        if org and contract and isinstance(org.widget, QComboBox):
            org.widget.currentTextChanged.connect(self._on_guarantee_org_changed)

    def _on_manager_changed(self, manager_name: str) -> None:
        value = self.cfg.manager_to_branch.get(clean_text(manager_name), "")
        if value and "branch_name" in self.field_boxes:
            self.field_boxes["branch_name"].set_value(value)

    def _on_guarantee_org_changed(self, org_name: str) -> None:
        value = self.cfg.guarantee_to_contract.get(clean_text(org_name), "")
        if value and "guarantee_contract_no" in self.field_boxes:
            self.field_boxes["guarantee_contract_no"].set_value(value)

    def get_field_value(self, key: str) -> str:
        fb = self.field_boxes.get(key)
        return fb.value() if fb else ""

    def collect_data(self) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
        data = {key: fb.value() for key, fb in self.field_boxes.items()}
        co_list: List[Dict[str, str]] = []
        for boxes in self.co_borrower_boxes:
            item = {key: fb.value() for key, fb in boxes.items()}
            # 全空的共同借款人行不导出
            if any(v for v in item.values()):
                co_list.append(item)
        return data, co_list

    def submit(self) -> None:
        data, co_list = self.collect_data()
        missing = []
        for fd in self._path_fields():
            if fd.required and fd.area != "共同借款人基本信息" and fd.key != "co_borrower_add":
                if not data.get(fd.key):
                    missing.append(fd.label)
        if missing:
            reply = QMessageBox.question(
                self,
                "存在未填必填项",
                "以下必填项尚未填写：\n" + "、".join(missing) + "\n\n是否仍然进入上传资料界面？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return
        self.submitted.emit(data, co_list, self.selected_path)


# =========================
# 页面：上传资料
# =========================

class AttachmentPage(QWidget):
    export_requested = Signal(dict, list, dict, str)  # data, co_list, attachments, selected_path
    back_requested = Signal()

    def __init__(self, cfg: AppConfig, template_dir: Path, form_data: Dict[str, str], co_list: List[Dict[str, str]], selected_path: str):
        super().__init__()
        self.cfg = cfg
        self.template_dir = template_dir
        self.form_data = form_data
        self.co_list = co_list
        self.selected_path = selected_path
        self.drop_areas: Dict[int, DropArea] = {}

        self.setStyleSheet("background: #E2F0D9;")
        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        top = QHBoxLayout()
        back_btn = QPushButton("返回修改信息")
        back_btn.setStyleSheet("font-size: 13px; padding: 6px 10px;")
        back_btn.clicked.connect(self.back_requested.emit)
        title = QLabel("提交资料")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: 700; color: #1f1f1f;")
        top.addWidget(back_btn)
        top.addWidget(title, 1)
        top.addSpacing(100)
        root.addLayout(top)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: #E2F0D9; }")
        body = QWidget()
        body.setStyleSheet("background: #E2F0D9;")
        grid = QGridLayout(body)
        grid.setContentsMargins(4, 4, 4, 4)
        grid.setHorizontalSpacing(0)
        grid.setVerticalSpacing(0)
        scroll.setWidget(body)
        root.addWidget(scroll, 1)

        header1 = self._table_cell("需要的文件名称", bold=True, center=True)
        header2 = self._table_cell("上传附件", bold=True, center=True)
        grid.addWidget(header1, 0, 0)
        grid.addWidget(header2, 0, 1)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 2)

        row_idx = 1
        for item in self.cfg.upload_items:
            if item.path_scope not in ("全部", "", self.selected_path):
                # 可扩展：配置里也可写“路径A”或“路径B”
                if "路径A" in self.selected_path and "路径A" not in item.path_scope:
                    continue
                if "路径B" in self.selected_path and "路径B" not in item.path_scope:
                    continue
            left = self._left_cell(item)
            drop = DropArea(item)
            self.drop_areas[item.row_no] = drop
            right_frame = QFrame()
            right_frame.setStyleSheet("QFrame { background: #ffffff; border: 1px solid #000; }")
            right_layout = QVBoxLayout(right_frame)
            right_layout.setContentsMargins(12, 12, 12, 12)
            right_layout.addWidget(drop)
            grid.addWidget(left, row_idx, 0)
            grid.addWidget(right_frame, row_idx, 1)
            row_idx += 1

        bottom = QHBoxLayout()
        bottom.addStretch(1)
        finish_btn = QPushButton("完成")
        finish_btn.setMinimumWidth(150)
        finish_btn.setFixedHeight(42)
        finish_btn.setCursor(Qt.PointingHandCursor)
        finish_btn.setStyleSheet("font-size: 16px; font-weight: 700; background: #2f6f3e; color: white; border-radius: 6px;")
        finish_btn.clicked.connect(self.finish)
        bottom.addWidget(finish_btn)
        root.addLayout(bottom)

    def _table_cell(self, text: str, bold: bool = False, center: bool = False) -> QLabel:
        label = QLabel(text)
        label.setWordWrap(True)
        label.setTextFormat(Qt.RichText)
        label.setStyleSheet(
            "QLabel { background: #ffffff; border: 1px solid #000; padding: 12px; "
            f"font-size: {'14' if bold else '13'}px; font-weight: {'700' if bold else '400'}; color: #111; }}"
        )
        if center:
            label.setAlignment(Qt.AlignCenter)
        else:
            label.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        return label

    def _left_cell(self, item: UploadItem) -> QLabel:
        text = item.name.replace("\n", "<br>")
        if item.template_label and item.template_prefix:
            text += f' <a href="template:{item.template_prefix}" style="color:#005bbb; text-decoration: underline;">{item.template_label}</a>'
        if item.required:
            text += ' <span style="color:#cc0000;">*</span>'
        label = self._table_cell(text)
        label.linkActivated.connect(self._open_template_link)
        return label

    def _open_template_link(self, href: str) -> None:
        if not href.startswith("template:"):
            return
        prefix = href.split(":", 1)[1]
        path = find_template_file(self.template_dir, prefix)
        if path:
            open_file(path)
        # 按需求：没有对应文件则无反应

    def finish(self) -> None:
        missing = []
        attachments: Dict[int, List[Path]] = {}
        for item in self.cfg.upload_items:
            area = self.drop_areas.get(item.row_no)
            files = list(area.files) if area else []
            attachments[item.row_no] = files
            if item.required and not files:
                missing.append(item.name)
        if missing:
            reply = QMessageBox.question(
                self,
                "存在未上传必传资料",
                "以下资料尚未上传：\n" + "\n".join(missing) + "\n\n是否仍然继续导出？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return

        QMessageBox.information(self, "即将导出", "将导出参数脚本及备案表及汇总文件夹。")
        folder = QFileDialog.getExistingDirectory(self, "选择保存路径")
        if not folder:
            return
        attachments = {k: v for k, v in attachments.items()}
        self.export_requested.emit(self.form_data, self.co_list, attachments, folder)


# =========================
# 导出逻辑
# =========================

class Exporter:
    def __init__(self, cfg: AppConfig, progress: Optional[QProgressDialog] = None):
        self.cfg = cfg
        self.progress = progress
        self.progress_value = 0

    def _tick(self, text: str, value: Optional[int] = None) -> None:
        if not self.progress:
            return
        if value is not None:
            self.progress_value = value
        else:
            self.progress_value = min(100, self.progress_value + 5)
        self.progress.setLabelText(text)
        self.progress.setValue(self.progress_value)
        QApplication.processEvents()

    def export_all(
        self,
        form_data: Dict[str, str],
        co_list: List[Dict[str, str]],
        attachments: Dict[int, List[Path]],
        save_dir: Path,
    ) -> Path:
        self._tick("正在清洗填写数据...", 5)
        clean_data = {k: clean_text(v) for k, v in form_data.items()}

        branch = safe_filename(clean_data.get("branch_name"))
        manager = safe_filename(clean_data.get("manager_name"))
        borrower = safe_filename(clean_data.get("borrower_name"))
        receipt = safe_filename(clean_data.get("loan_receipt_no"))
        base_name = safe_filename(f"{branch}-{manager}-{borrower}-{receipt}")

        # 避免所有导出散落在同一路径：创建一个汇总文件夹
        output_root = save_dir / base_name
        output_root.mkdir(parents=True, exist_ok=True)

        self._tick("正在生成备案Excel...", 15)
        excel_path = output_root / f"{base_name}.xlsx"
        self._write_export_excel(excel_path, clean_data, co_list)

        self._tick("正在创建资料归档文件夹...", 45)
        self._copy_attachments(output_root, attachments)

        self._tick("正在生成参数脚本...", 90)
        self._write_param_script(output_root, clean_data, co_list)

        self._tick("导出完成。", 100)
        return output_root

    def _value_for_col(self, col: ExportCol, data: Dict[str, str], co_list: List[Dict[str, str]]) -> Any:
        source = col.source_key
        if not source:
            return self._typed_value(col.default, col)

        if "取借据起始日月份" in col.rule or col.col_name == "所属月份":
            return extract_month(data.get(source, ""))

        if source in {"loan_start_date", "loan_end_date"} or "规范日期" in col.rule:
            return normalize_date(data.get(source, ""))

        if source.startswith("co_borrower"):
            values = [clean_text(x.get(source, "")) for x in co_list if clean_text(x.get(source, ""))]
            return "/".join(values)

        value = data.get(source, "")
        if is_blank(value):
            value = col.default
        return self._typed_value(value, col)

    def _typed_value(self, value: Any, col: ExportCol) -> Any:
        if is_blank(value):
            return ""
        data_type = col.data_type
        if any(x in data_type for x in ["金额", "数字", "整数", "百分比"]):
            num = to_number(value)
            if num is None:
                return clean_text(value)
            if "整数" in data_type:
                return int(num)
            return num
        if "日期" in data_type and "文本/日期" not in data_type:
            return normalize_date(value)
        return clean_text(value)

    def _write_export_excel(self, excel_path: Path, data: Dict[str, str], co_list: List[Dict[str, str]]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "备案表"

        headers = [c.col_name for c in self.cfg.export_cols]
        ws.append(headers)
        row_values = [self._value_for_col(c, data, co_list) for c in self.cfg.export_cols]
        ws.append(row_values)

        header_fill = PatternFill("solid", fgColor="D9EAD3")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = Font(name="微软雅黑", size=10, bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                cell.font = Font(name="微软雅黑", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = border

        for idx, col in enumerate(self.cfg.export_cols, start=1):
            cell = ws.cell(row=2, column=idx)
            if col.excel_format:
                cell.number_format = col.excel_format
            # 文本格式强制写字符串，避免身份证号/合同号科学计数法
            if col.excel_format == "@" and cell.value is not None:
                cell.value = clean_text(cell.value)

        # 自动列宽
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 10
            for cell in ws[letter]:
                text = as_text(cell.value)
                max_len = max(max_len, min(30, len(text) + 2))
            ws.column_dimensions[letter].width = max_len
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"
        wb.save(excel_path)

    def _copy_attachments(self, output_root: Path, attachments: Dict[int, List[Path]]) -> None:
        upload_by_row = {u.row_no: u for u in self.cfg.upload_items}
        total_rows = max(1, len(upload_by_row))
        done = 0
        for row_no, item in upload_by_row.items():
            folder_name = safe_filename(item.name.replace("\n", " "))
            dest_dir = output_root / f"{row_no:02d}.{folder_name}"
            dest_dir.mkdir(parents=True, exist_ok=True)
            for src in attachments.get(row_no, []):
                if not src.exists() or not src.is_file():
                    continue
                dest = dest_dir / safe_filename(src.name)
                dest = self._unique_path(dest)
                shutil.copy2(src, dest)
            done += 1
            self._tick(f"正在归档附件：{item.name}", 45 + int(done / total_rows * 40))

    def _unique_path(self, path: Path) -> Path:
        if not path.exists():
            return path
        stem = path.stem
        suffix = path.suffix
        parent = path.parent
        i = 1
        while True:
            candidate = parent / f"{stem}_{i}{suffix}"
            if not candidate.exists():
                return candidate
            i += 1

    def _write_param_script(self, output_root: Path, data: Dict[str, str], co_list: List[Dict[str, str]]) -> None:
        """生成一个简单参数脚本/摘要，方便排查导出时使用的参数。"""
        script_path = output_root / "导出参数脚本.txt"
        lines = []
        lines.append("资料分类汇总程序 - 导出参数脚本/摘要")
        lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"配置文件：{self.cfg.config_path}")
        lines.append("")
        lines.append("[填写信息]")
        for k in sorted(data):
            lines.append(f"{k}={data[k]}")
        lines.append("")
        lines.append("[共同借款人]")
        if co_list:
            for idx, item in enumerate(co_list, start=1):
                lines.append(f"共同借款人{idx}：" + "; ".join(f"{k}={v}" for k, v in item.items()))
        else:
            lines.append("无")
        script_path.write_text("\n".join(lines), encoding="utf-8")


# =========================
# 主窗口
# =========================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setMinimumSize(1180, 760)
        self.cfg: Optional[AppConfig] = None
        self.template_dir: Path = Path("")
        self.current_form: Optional[FormPage] = None
        self.current_attachment_page: Optional[AttachmentPage] = None

        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        self.startup_page = StartupPage()
        self.startup_page.loaded.connect(self.on_config_loaded)
        self.stack.addWidget(self.startup_page)
        self.setWindowTitle("资料分类汇总程序")

    def on_config_loaded(self, cfg: AppConfig, template_dir: Path) -> None:
        self.cfg = cfg
        self.template_dir = template_dir
        self.setWindowTitle(cfg.program_name)
        self.path_page = PathSelectPage(cfg)
        self.path_page.path_selected.connect(self.show_form_page)
        self.stack.addWidget(self.path_page)
        self.stack.setCurrentWidget(self.path_page)

    def show_form_page(self, path_name: str) -> None:
        assert self.cfg is not None
        if self.current_form:
            self.stack.removeWidget(self.current_form)
            self.current_form.deleteLater()
        self.current_form = FormPage(self.cfg, path_name)
        self.current_form.back_requested.connect(lambda: self.stack.setCurrentWidget(self.path_page))
        self.current_form.submitted.connect(self.show_attachment_page)
        self.stack.addWidget(self.current_form)
        self.stack.setCurrentWidget(self.current_form)

    def show_attachment_page(self, form_data: Dict[str, str], co_list: List[Dict[str, str]], selected_path: str) -> None:
        assert self.cfg is not None
        if self.current_attachment_page:
            self.stack.removeWidget(self.current_attachment_page)
            self.current_attachment_page.deleteLater()
        self.current_attachment_page = AttachmentPage(self.cfg, self.template_dir, form_data, co_list, selected_path)
        self.current_attachment_page.back_requested.connect(lambda: self.stack.setCurrentWidget(self.current_form))
        self.current_attachment_page.export_requested.connect(self.export_all)
        self.stack.addWidget(self.current_attachment_page)
        self.stack.setCurrentWidget(self.current_attachment_page)

    def export_all(
        self,
        form_data: Dict[str, str],
        co_list: List[Dict[str, str]],
        attachments: Dict[int, List[Path]],
        folder: str,
    ) -> None:
        assert self.cfg is not None
        progress = QProgressDialog("准备导出...", "取消", 0, 100, self)
        progress.setWindowTitle("导出中")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        progress.show()
        QApplication.processEvents()

        try:
            exporter = Exporter(self.cfg, progress)
            output_root = exporter.export_all(form_data, co_list, attachments, Path(folder))
            QMessageBox.information(self, "导出完成", f"资料已导出至：\n{output_root}")
            # 按需求：完成后自动关闭整个程序
            QApplication.quit()
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"导出过程中发生错误：\n{exc}\n\n{traceback.format_exc()}")
        finally:
            progress.close()


def main() -> int:
    app = QApplication(sys.argv)
    app.setFont(QFont("微软雅黑", 10))
    win = MainWindow()
    win.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
